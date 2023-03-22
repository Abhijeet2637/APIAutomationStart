import requests
import json
import jsonpath
import openpyxl

def test_add_multiple_student():
    URL= "https://thetestingworldapi.com/api/studentsDetails"
    file=open('C:/Users/user/Desktop/StudentDetails/StudentsDetails.json')
    json_file=json.loads(file.read())
    #open Excel
    wk=openpyxl.load_workbook('C:/Users/user/Desktop/StudentDetails/Multiple.xlsx')
    sh=wk['Sheet1']
    rows=sh.max_row

    #retrieve excel data
    for i in range(2,rows+1):
        cell_first_name=sh.cell(row=i, column=1)
        cell_mid_name=sh.cell(row=i,column=2)
        cell_last_name=sh.cell(row=i,column=3)
        cell_dob=sh.cell(row=i,column=4)

        #use values
        json_file['first_name']=cell_first_name.value
        json_file['middle_name']=cell_mid_name.value
        json_file['last_name']=cell_last_name.value
        json_file['date_of_birth']=cell_dob.value

        #send request
        response=requests.post(URL,json_file)
        print(response.text)
        print(response.status_code)

        #verification
        assert response.status_code==201








